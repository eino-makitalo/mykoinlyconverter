#--*-- coding: utf-8 --*--

import csv
import datetime
import pytz
import shutil
import os
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter

hmap={}

WORKDIR=r"c:\temp"
from settings import WORKDIR  # settings.py is inl

FIFO_TEMPLATE="verohallinto_-fifo-laskuri_versio-1.1.xlsm" 

ASSET_COL_IN_TEMPLATE=8   # column H  after text "Laskelmassa käytetty virtuaalivaluutta"
ASSET_ROW_IN_TEMPLATE=9   # row where currency is saved (just information)

TIMESCALE_COL_IN_TEMPLATE=1
TIMESCALE_ROW_IN_TEMPLATE=4

FIFO_START_ROW=12

FIFO_COL_DATE=1 #AIKA
FIFO_COL_EVTYPE=2 #TAPAHTUMA "Osto" / "Myynti"
FIFO_COL_AMOUNT=3 #MÄÄRÄ / Amount in asset
FIFO_COL_PRICE=4 #HINTA PER VIR..   price in Euros of one crypto asset
FIFO_COL_EUROS=5 #YHTEENSÄ
FIFO_COL_SOURCE=6




KOINLYTRANSACTIONFILES_TO_READ = ("Koinly_2021.csv","Koinly_2022.csv")


def ExcelForCurrency(currency):
    return "%s_FIFO.xlsm" % (currency,)

hki = pytz.timezone("Europe/Helsinki")

BOOK_CURRENCY="EUR"  # bookkeeping currency
KNOWN_CURRENCIES=["EUR","USD"]

F_DATE="Date"
F_TYPE="Type"
F_LABEL="Label"
F_SWALLET="Sending Wallet"
F_SAMOUNT="Sent Amount" # in crypto
F_SCURRENCY="Sent Currency"
F_SCOST="Sent Cost Basis" # in bookkeeping currency
F_RWALLET="Receiving Wallet"
F_RAMOUNT="Received Amount" # in crypto
F_RCURRENCY="Received Currency"
F_RCOST="Received Cost Basis" # in bookkeeping currency
F_FEE="Fee Amount"
F_FCURRENCY="Fee Currency"
F_GAIN="Gain (%s)" % (BOOK_CURRENCY,)
F_NETVAL="Net Value (%s)" % (BOOK_CURRENCY,)
F_FEEVAK="Fee Value (%s)" % (BOOK_CURRENCY,)
F_TXSRC="TxSrc"
F_TXDEST="TxDest"
F_TXHASH="TxHash"
F_DESC="Description"

# added by use
F_LOCALTIME="Localtime"

# Date,Type,Label,Sending Wallet,Sent Amount,
# Sent Currency,Sent Cost Basis,Receiving Wallet,Received Amount,Received Currency,
# Received Cost Basis,Fee Amount,Fee Currency,
# Gain (EUR),Net Value (EUR),Fee Value (EUR),TxSrc,TxDest,TxHash,Description

TIMEZONE="Europe/Helsinki"  # just used to separate months for bookkeeping

FIFO_ABS=os.path.join(WORKDIR,FIFO_TEMPLATE)
if not os.path.exists(FIFO_ABS):
    raise AssertionError("""We cant file file %s  
    in directory %s
    download it from\n https://www.vero.fi/tietoa-verohallinnosta/yhteystiedot-ja-asiointi/verohallinnon_laskuri/fifo-laskuri/"""%(FIFO_TEMPLATE,WORKDIR))

import os

TYPES=set()
Currencies=set()
Wallets=set()

ROWS=[]

first_csv_file=True
startdate=enddate=None
for file1 in KOINLYTRANSACTIONFILES_TO_READ:
    absfile1=os.path.join(WORKDIR,file1)
    skip_until_header=True
    with open(absfile1,"r") as csvfile:
        for row in csv.reader(csvfile,dialect='excel'):        
            if skip_until_header:
                if len(row)==20 and row[0]=='Date':
                    skip_until_header=False
                    # we read this from first csv only - should be same mapping
                    ind=0
                    for cell in row:
                        if first_csv_file:
                            hmap[cell]=ind
                        else:
                            if (ind != hmap[cell]):
                                raise AssertionError("Transaction files from Koinly has different format - can't continue")
                        ind=ind+1  
                        hmap[F_LOCALTIME]=ind  # we add also local time here
            else:
                utctime=pytz.utc.localize(datetime.datetime.strptime(row[0],'%Y-%m-%d %H:%M:%S UTC'))
                #print(utctime)
                x1=hki.normalize(utctime)
                if not startdate:
                    startdate=x1
                enddate=x1
                print(utctime, x1)
                tstamp=str(x1)[:19]
                #datepart,timepart=tstamp.split(" ")
                #yy,mm,dd=datepart.split("-")
                #datepart2=".".join((dd,mm,yy))
                row.append(x1.replace(tzinfo=None))
                TYPES.add(row[hmap[F_TYPE]])
                if row[hmap[F_SCURRENCY]]:
                    Currencies.add(row[hmap[F_SCURRENCY]])
                if row[hmap[F_RCURRENCY]]:
                    Currencies.add(row[hmap[F_RCURRENCY]])
                if row[hmap[F_RWALLET]]:
                    Wallets.add(row[hmap[F_RWALLET]])
                if row[hmap[F_SWALLET]]:
                    Wallets.add(row[hmap[F_SWALLET]])
                ROWS.append(row)
# We will create one Tax calculation Excel for every wallet
# Please, if you have same currency USDC in multiple blockchain you should consider handling them separately ?


TIMESCALES=[]
start_year=startdate.year
end_year=enddate.year
year=start_year
while(year<=end_year):
    TIMESCALES.append("1.1.%s-31.12.%s" % (str(year),str(year)))
    year=year+1
                      

print("Time scale ",TIMESCALES)

EXCELMAP={}

import tempfile

def fixDes(summa):
    return float(summa)

def saveexcel(wb,absfile):
    dir1,file1=os.path.split(absfile)
    fn1,ext1 = os.path.splitext(file1)    
    fdtemp,absfiletemp=tempfile.mkstemp(suffix=ext1,prefix=fn1)
    os.close(fdtemp)
    wb.save(absfiletemp)
    shutil.copy2(absfiletemp,absfile)
    os.remove(absfiletemp)
    

for curr in Currencies:    
    if curr in KNOWN_CURRENCIES:
        continue
    absfile2=os.path.join(WORKDIR,ExcelForCurrency(curr))
    EXCELMAP[curr]=absfile2
    if os.path.exists(absfile2):
        print(f"We have already excel for {curr}")
    else:
        print(f"We create new exec for {curr} from template")
        shutil.copy2(os.path.join(WORKDIR,FIFO_TEMPLATE), absfile2)
    wb=load_workbook(absfile2,keep_vba=True)
    first_sheet=wb.worksheets[0]
    curr2=first_sheet.cell(ASSET_ROW_IN_TEMPLATE,ASSET_COL_IN_TEMPLATE).value
    if (curr!=curr2):
        assert(curr2==None)
        print(f"We set excel right currency {curr}")
        first_sheet.cell(ASSET_ROW_IN_TEMPLATE,ASSET_COL_IN_TEMPLATE).value=curr
    ind=0
    for t in TIMESCALES:
        first_sheet.cell(TIMESCALE_ROW_IN_TEMPLATE+ind,TIMESCALE_COL_IN_TEMPLATE).value=t
        ind=ind+1
    # handle all transactions
    rowind=FIFO_START_ROW
    for row in ROWS:
        if (row[hmap[F_SCURRENCY]]==curr) or (row[hmap[F_RCURRENCY]]==curr):
            first_sheet.cell(rowind,FIFO_COL_DATE).value=row[hmap[F_LOCALTIME]]
            maintype=row[hmap[F_TYPE]]
            extinfo=row[hmap[F_LABEL]]
            
            if maintype=="crypto_deposit":
                first_sheet.cell(rowind,FIFO_COL_EVTYPE).value="Osto"
                first_sheet.cell(rowind,FIFO_COL_AMOUNT).value=fixDes(row[hmap[F_RAMOUNT]])
                first_sheet.cell(rowind,FIFO_COL_EUROS).value=fixDes(row[hmap[F_NETVAL]])
                first_sheet.cell(rowind,FIFO_COL_PRICE).value=fixDes(row[hmap[F_NETVAL]])/fixDes(row[hmap[F_RAMOUNT]])
                first_sheet.cell(rowind,FIFO_COL_SOURCE).value=row[hmap[F_RWALLET]]
                
            elif maintype=="crypto_withdrawal":
                first_sheet.cell(rowind,FIFO_COL_EVTYPE).value="Myynti"
                first_sheet.cell(rowind,FIFO_COL_AMOUNT).value=fixDes(row[hmap[F_SAMOUNT]])
                first_sheet.cell(rowind,FIFO_COL_EUROS).value=fixDes(row[hmap[F_NETVAL]])
                first_sheet.cell(rowind,FIFO_COL_PRICE).value=fixDes(row[hmap[F_NETVAL]])/fixDes(row[hmap[F_SAMOUNT]])
                first_sheet.cell(rowind,FIFO_COL_SOURCE).value=row[hmap[F_SWALLET]]
            elif maintype=="buy":
                #assert(row[hmap[F_SCURRENCY]]==BOOK_CURRENCY) # hmm... you can use also dollars so...
                #assert(row[hmap[F_SAMOUNT]]==row[hmap[F_NETVAL]])
                first_sheet.cell(rowind,FIFO_COL_EVTYPE).value="Osto"
                first_sheet.cell(rowind,FIFO_COL_AMOUNT).value=fixDes(row[hmap[F_RAMOUNT]])
                first_sheet.cell(rowind,FIFO_COL_EUROS).value=fixDes(row[hmap[F_NETVAL]])
                first_sheet.cell(rowind,FIFO_COL_PRICE).value=fixDes(row[hmap[F_NETVAL]])/fixDes(row[hmap[F_RAMOUNT]])
                first_sheet.cell(rowind,FIFO_COL_SOURCE).value=row[hmap[F_RWALLET]]
            elif maintype=="exchange":
                if (row[hmap[F_FEE]]!='') and (row[hmap[F_RCURRENCY]]==curr):
                    fee=fixDes(row[hmap[F_FEE]])
                else:
                    fee=0
                if row[hmap[F_SCURRENCY]]==curr: #Sent currency is current one so its sell / Myynti
                    first_sheet.cell(rowind,FIFO_COL_EVTYPE).value="Myynti"
                    first_sheet.cell(rowind,FIFO_COL_AMOUNT).value=fixDes(row[hmap[F_SAMOUNT]])                    
                    first_sheet.cell(rowind,FIFO_COL_SOURCE).value=row[hmap[F_SWALLET]]
                    first_sheet.cell(rowind,FIFO_COL_EUROS).value=fixDes(row[hmap[F_NETVAL]])   
                    first_sheet.cell(rowind,FIFO_COL_PRICE).value=fixDes(row[hmap[F_NETVAL]])/fixDes(row[hmap[F_SAMOUNT]])
                    
                else:
                    first_sheet.cell(rowind,FIFO_COL_EVTYPE).value="Osto"
                    first_sheet.cell(rowind,FIFO_COL_AMOUNT).value=fixDes(row[hmap[F_RAMOUNT]])-fee
                    first_sheet.cell(rowind,FIFO_COL_SOURCE).value=row[hmap[F_RWALLET]]
                    first_sheet.cell(rowind,FIFO_COL_EUROS).value=fixDes(row[hmap[F_NETVAL]])      
                    first_sheet.cell(rowind,FIFO_COL_PRICE).value=fixDes(row[hmap[F_NETVAL]])/fixDes(row[hmap[F_RAMOUNT]])
                    
            elif maintype=="sell":
                #assert(row[hmap[F_RCURRENCY]]==BOOK_CURRENCY)
                #assert(row[hmap[F_SAMOUNT]]==row[hmap[F_NETVAL]])
                first_sheet.cell(rowind,FIFO_COL_EVTYPE).value="Myynti"
                first_sheet.cell(rowind,FIFO_COL_AMOUNT).value=fixDes(row[hmap[F_SAMOUNT]])
                first_sheet.cell(rowind,FIFO_COL_EUROS).value=fixDes(row[hmap[F_NETVAL]])
                first_sheet.cell(rowind,FIFO_COL_SOURCE).value=row[hmap[F_SWALLET]]
                first_sheet.cell(rowind,FIFO_COL_PRICE).value=fixDes(row[hmap[F_NETVAL]])/fixDes(row[hmap[F_SAMOUNT]])
                
            elif maintype=="transfer":
                #assert(row[hmap[F_RCURRENCY]]==BOOK_CURRENCY)
                #assert(row[hmap[F_SAMOUNT]]==row[hmap[F_NETVAL]])
                # we handle possible transfer fees
                if row[hmap[F_FEE]]:
                    assert(row[hmap[F_FCURRENCY]]==curr)
                    first_sheet.cell(rowind,FIFO_COL_EVTYPE).value="Myynti"
                    first_sheet.cell(rowind,FIFO_COL_AMOUNT).value=fixDes(row[hmap[F_FEE]])
                    first_sheet.cell(rowind,FIFO_COL_EUROS).value=fixDes(row[hmap[F_FEEVAK]])
                    first_sheet.cell(rowind,FIFO_COL_SOURCE).value=row[hmap[F_SWALLET]]
                    first_sheet.cell(rowind,FIFO_COL_PRICE).value=fixDes(row[hmap[F_FEEVAK]])/fixDes(row[hmap[F_FEE]])
                    
            else:
                print("skip row",maintype,row)
            pass
            rowind=rowind+1
        
    saveexcel(wb, absfile2)
    print(f"Currency {curr} is set in excel workbook")
        
        

print(TYPES)
print(Currencies)
print(Wallets)